#!/usr/bin/env python3
"""Backfill missing parameters_m for models from interpolation Excels/zips."""
import json
import zipfile
import tempfile
from pathlib import Path

ROOT = Path(r'C:\Code\benchmark')
MODELS_PATH = ROOT / 'src' / 'data' / 'models.json'

SEGC3_EXCEL = Path(r'C:\Users\admin\Documents\WeChat Files\wxid_hvmr1h95e7jn22\FileStorage\File\2026-08\batch_evaluation.xlsx')
AVO_EXCEL = Path(r'C:\Users\admin\Documents\WeChat Files\wxid_hvmr1h95e7jn22\FileStorage\File\2026-08\batch_evaluation_avo.xlsx')
SEGC3_ZIP = Path(r'C:\Users\admin\Documents\WeChat Files\wxid_hvmr1h95e7jn22\FileStorage\File\2026-08\interp_syn_czt0822.zip')
AVO_ZIP = Path(r'C:\Users\admin\Documents\WeChat Files\wxid_hvmr1h95e7jn22\FileStorage\File\2026-08\avo_interpolation_json.zip')
FIELD_ZIP = Path(r'C:\Users\admin\Documents\WeChat Files\wxid_hvmr1h95e7jn22\FileStorage\File\2026-08\interp_field_czt0820.zip')

# Map Excel method name -> repo model IDs that should receive this parameter count
METHOD_TO_MODELS = {
    'UNet': ['unet-interpolation', 'avo-unet-interpolation'],
    'UNet (UNPN)': ['unet-unpn-interpolation', 'avo-unet-unpn-interpolation'],
    'UNet-Plus': ['unet-plus-interpolation', 'avo-unet-plus-interpolation'],
    'ResUNet': ['res-unet-interpolation', 'avo-resunet-interpolation'],
    'ResUNet(UNPN)': ['resunet-unpn-interpolation', 'avo-resunet-unpn-interpolation'],
    'ResUNet-Plus': ['resunet-plus-interpolation', 'avo-resunet-plus-interpolation'],
    'Attention UNet': ['attention-unet-interpolation', 'avo-attention-unet-interpolation'],
    'Attention UNet(UNPN)': ['attention-unet-unpn-interpolation', 'avo-attention-unet-unpn-interpolation'],
    'Attention UNet-Plus': ['attention-unet-plus-interpolation', 'avo-attention-unet-plus-interpolation'],
    'DnCNN': ['dncnn-interpolation', 'avo-dncnn-interpolation'],
    'DnCNN-Plus': ['dncnn-plus-interpolation', 'avo-dncnn-plus-interpolation'],
    'SPNet': ['spnet-interpolation', 'avo-spnet-interpolation'],
}


def load_json(path):
    with path.open('r', encoding='utf-8') as f:
        return json.load(f)


def save_json(path, data):
    with path.open('w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write('\n')


def parse_excel_params(path):
    import openpyxl
    params = {}
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c).strip() if c else '' for c in rows[0]]
        try:
            method_idx = headers.index('Method')
            param_idx = headers.index('Parameters (M)')
        except ValueError:
            continue
        for row in rows[1:]:
            method = str(row[method_idx]).strip() if row[method_idx] else ''
            val = row[param_idx]
            if method and val is not None:
                # keep first encountered (same across sheets)
                params.setdefault(method, float(val))
    return params


def apply_excel_params(models, excel_path):
    params = parse_excel_params(excel_path)
    updated = 0
    for method, val in params.items():
        for mid in METHOD_TO_MODELS.get(method, []):
            for m in models:
                if m['id'] == mid and m.get('parameters_m') is None:
                    m['parameters_m'] = val
                    updated += 1
                    print(f'Updated {mid}: {val} M')
    return updated


def extract_zip_models(zippath):
    """Yield (raw_id, parameters_m) from model JSONs in zip."""
    tmp = tempfile.mkdtemp()
    with zipfile.ZipFile(zippath, 'r') as z:
        z.extractall(tmp)
    root = Path(tmp)
    for p in root.rglob('*.json'):
        if '__MACOSX' in p.parts:
            continue
        name = p.name.lower()
        if 'model' in name and 'detailed' not in name:
            try:
                data = load_json(p)
            except Exception:
                continue
            mid = data.get('id')
            param = data.get('parameters_m')
            if mid and param is not None:
                yield mid, float(param)


def apply_zip_params(models, zippath):
    updated = 0
    by_id = {m['id']: m for m in models}
    for raw_id, val in extract_zip_models(zippath):
        targets = [raw_id]
        if raw_id + '_interpolation' in by_id:
            targets.append(raw_id + '_interpolation')
        # avo- prefix variants
        if raw_id.startswith('avo-'):
            targets.append(raw_id)
        else:
            avo_id = 'avo-' + raw_id
            if avo_id in by_id:
                targets.append(avo_id)
            avo_interp = 'avo-' + raw_id + '_interpolation'
            if avo_interp in by_id:
                targets.append(avo_interp)
        for mid in set(targets):
            m = by_id.get(mid)
            if m and m.get('parameters_m') is None:
                m['parameters_m'] = val
                updated += 1
                print(f'Updated {mid}: {val} M')
    return updated


def main():
    models = load_json(MODELS_PATH)
    before = sum(1 for m in models if m.get('parameters_m') is None)
    print(f'Models missing parameters_m before: {before}')

    total = 0
    total += apply_excel_params(models, SEGC3_EXCEL)
    total += apply_excel_params(models, AVO_EXCEL)
    for zp in [SEGC3_ZIP, AVO_ZIP, FIELD_ZIP]:
        total += apply_zip_params(models, zp)

    after = sum(1 for m in models if m.get('parameters_m') is None)
    print(f'Updated {total} entries; missing after: {after}')

    save_json(MODELS_PATH, models)
    print('Done')


if __name__ == '__main__':
    main()
