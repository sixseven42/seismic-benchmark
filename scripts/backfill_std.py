#!/usr/bin/env python3
"""Backfill scores_std for existing results from source Excels/zips."""
import json
import re
import zipfile
import tempfile
from pathlib import Path
from collections import defaultdict

ROOT = Path(r'C:\Code\benchmark')
RESULTS_PATH = ROOT / 'src' / 'data' / 'results.json'

SEGC3_INTERP_ZIP = Path(r'C:\Users\admin\Documents\WeChat Files\wxid_hvmr1h95e7jn22\FileStorage\File\2026-08\interpolation_json.zip')
AVO_INTERP_ZIP = Path(r'C:\Users\admin\Documents\WeChat Files\wxid_hvmr1h95e7jn22\FileStorage\File\2026-08\avo_interpolation_json.zip')
UNET_PP_EXCEL = Path(r'C:\论文\SeisBench\seismic_bench可视化\random\batch_evaluation_unet_plusplus.xlsx')
GROUND_ROLL_EXCEL = Path(r'C:\Users\admin\Documents\WeChat Files\wxid_hvmr1h95e7jn22\FileStorage\File\2026-07\batch_evaluation_all.xlsx')
MULTIPLES_EXCEL = Path(r'C:\Users\admin\Documents\WeChat Files\wxid_hvmr1h95e7jn22\FileStorage\File\2026-07\batch_evaluation_part(2).xlsx')

CORE_METRICS = ['snr', 'psnr', 'ssim', 'mae', 'mse', 'rmse']
BINNED_METRICS = [
    'eb_wse_medium_40_70_ne', 'eb_wse_medium_40_70_snr',
    'eb_wse_strong_70_100_ne', 'eb_wse_strong_70_100_snr',
    'eb_wse_very_weak_5_20_ne', 'eb_wse_very_weak_5_20_snr',
    'eb_wse_weak_20_40_ne', 'eb_wse_weak_20_40_snr',
    'fb_fre_high_ne', 'fb_fre_high_snr',
    'fb_fre_low_ne', 'fb_fre_low_snr',
    'fb_fre_mid_ne', 'fb_fre_mid_snr',
    'fb_fre_very_high_ne', 'fb_fre_very_high_snr',
]
VALID_METRICS = set(CORE_METRICS + BINNED_METRICS)

SEGC3_MODEL_MAP = {
    'unet-interpolation': 'unet-interpolation',
    'unet-unpn-interpolation': 'unet-unpn-interpolation',
    'unet-plus-interpolation': 'unet-plus-interpolation',
    'attention-unet-interpolation': 'attention-unet-interpolation',
    'attention-unet-unpn-interpolation': 'attention-unet-unpn-interpolation',
    'attention-unet-plus-interpolation': 'attention-unet-plus-interpolation',
    'resunet-interpolation': 'res-unet-interpolation',
    'resunet-unpn-interpolation': 'resunet-unpn-interpolation',
    'resunet-plus-interpolation': 'resunet-plus-interpolation',
    'dncnn-interpolation': 'dncnn-interpolation',
    'dncnn-plus-interpolation': 'dncnn-plus-interpolation',
    'spnet-interpolation': 'spnet-interpolation',
}

AVO_MODEL_MAP = {mid: mid for mid in [
    'avo-unet-interpolation',
    'avo-unet-plus-interpolation',
    'avo-attention-unet-interpolation',
    'avo-attention-unet-plus-interpolation',
    'avo-resunet-interpolation',
    'avo-resunet-plus-interpolation',
    'avo-dncnn-interpolation',
    'avo-dncnn-plus-interpolation',
    'avo-spnet-interpolation',
]}

SEGC3_BENCH_MAP = {
    'interp-random30': 'segc3-interp-random30',
    'interp-random50': 'segc3-interp-random50',
    'interp-random70': 'segc3-interp-random70',
    'interp-uniform50': 'segc3-interp-uniform50',
    'interp-uniform75': 'segc3-interp-uniform75',
    'interp-continuous20tr': 'segc3-interp-continuous20tr',
    'interp-continuous30tr': 'segc3-interp-continuous30tr',
    'interp-continuous40tr': 'segc3-interp-continuous40tr',
}

AVO_BENCH_MAP = {
    'avo-interp-random30': 'mobile-avo-interp-random30',
    'avo-interp-random50': 'mobile-avo-interp-random50',
    'avo-interp-random70': 'mobile-avo-interp-random70',
    'avo-interp-uniform50': 'mobile-avo-interp-uniform50',
    'avo-interp-uniform75': 'mobile-avo-interp-uniform75',
    'avo-interp-continuous20tr': 'mobile-avo-interp-continuous20tr',
    'avo-interp-continuous30tr': 'mobile-avo-interp-continuous30tr',
    'avo-interp-continuous40tr': 'mobile-avo-interp-continuous40tr',
}

UNET_PP_SHEET_MAP = {
    'RN-SEGC3 Gaussian -5dB': 'segc3-random-noise-gaussian-snrneg5',
    'RN-SEGC3 Gaussian +0dB': 'segc3-random-noise-gaussian-snr0',
    'RN-SEGC3 Gaussian +5dB': 'segc3-random-noise-gaussian-snr5',
    'RN-SEGC3 Poisson -5dB': 'segc3-random-noise-poisson-snrneg5',
    'RN-SEGC3 Poisson +0dB': 'segc3-random-noise-poisson-snr0',
    'RN-SEGC3 Poisson +5dB': 'segc3-random-noise-poisson-snr5',
    'RN-AVO Gaussian -5dB': 'mobile-avo-random-noise-gaussian-snrneg5',
    'RN-AVO Gaussian +0dB': 'mobile-avo-random-noise-gaussian-snr0',
    'RN-AVO Gaussian +5dB': 'mobile-avo-random-noise-gaussian-snr5',
    'RN-AVO Poisson -5dB': 'mobile-avo-random-noise-poisson-snrneg5',
    'RN-AVO Poisson +0dB': 'mobile-avo-random-noise-poisson-snr0',
    'RN-AVO Poisson +5dB': 'mobile-avo-random-noise-poisson-snr5',
    'Deblending-AVO T03 mod': 'mobile-avo-deblending-t03-mod',
    'Deblending T02 mod': 'segc3-deblending-t02-mod',
    'Deblending T02 simp': 'segc3-deblending-t02-simp',
    'Deblending T02 comp': 'segc3-deblending-t02-comp',
}

GROUND_ROLL_METHOD_MAP = {
    'UNet': 'unet-groundroll',
    'UNet-Plus': 'unet-plus-groundroll',
    'ResUNet': 'res-unet-groundroll',
    'ResUNet-Plus': 'res-unet-plus-groundroll',
    'DnCNN': 'dncnn-groundroll',
    'Attention UNet': 'attention-unet-groundroll',
    'Attention UNet-Plus': 'attention-unet-plus-groundroll',
    'Enhanced Atten-UNet': 'enhanced-atten-unet-groundroll',
    'SANet': 'sanet-groundroll',
    'Physics CNN': 'physics-cnn-groundroll',
    'Pix2Pix cGAN': 'pix2pix-cgan-groundroll',
    'DDPM cDDPM': 'cddpm-groundroll',
}

GROUND_ROLL_SHEET_MAP = {
    'Noise 1.0': 'segc3-groundroll-noise1',
    'Noise 3.0': 'segc3-groundroll-noise3',
    'Noise 5.0': 'segc3-groundroll-noise5',
    'Noise 7.0': 'segc3-groundroll-noise7',
    'Noise 9.0': 'segc3-groundroll-noise9',
}

MULTIPLES_METHOD_MAP = {
    'UNet': 'unet-multiples',
    'UNet-Plus': 'unet-plus-multiples',
    'ResUNet': 'res-unet-multiples',
    'ResUNet-Plus': 'res-unet-plus-multiples',
    'DnCNN': 'dncnn-multiples',
    'Attention UNet': 'attention-unet-multiples',
    'Attention UNet-Plus': 'attention-unet-plus-multiples',
    'SAGAN': 'sagan-multiples',
    'DNNDAT': 'dnndat-multiples',
}


def load_json(path):
    with path.open('r', encoding='utf-8') as f:
        return json.load(f)


def save_json(path, data):
    with path.open('w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write('\n')


def parse_mean_std(value):
    """Return (mean, std) from a cell string like '9.2260±0.1077'."""
    if value is None:
        return None, None
    s = str(value).strip()
    if not s or s == '—':
        return None, None
    parts = re.split(r'[\s\xb1\u00b1]+', s)
    try:
        mean = float(parts[0])
    except ValueError:
        mean = None
    std = None
    if len(parts) > 1:
        try:
            std = float(parts[1])
        except ValueError:
            std = None
    return mean, std


def set_std(result, std_scores):
    if std_scores:
        result['scores_std'] = std_scores


def update_from_interpolation_zip(results_by_key, zippath, model_map, bench_map):
    updated = 0
    tmp = tempfile.mkdtemp()
    with zipfile.ZipFile(zippath, 'r') as z:
        z.extractall(tmp)
    root = Path(tmp)
    for p in root.rglob('*.json'):
        if '__MACOSX' in p.parts or 'detailed' in p.name.lower():
            continue
        if not ('result' in p.name.lower()):
            continue
        entries = load_json(p)
        if not isinstance(entries, list):
            entries = entries.get('results', [entries])
        for entry in entries:
            raw_bench = entry.get('benchmark_id')
            raw_model = entry.get('model_id')
            if raw_bench not in bench_map or raw_model not in model_map:
                continue
            bench_id = bench_map[raw_bench]
            model_id = model_map[raw_model]
            std = entry.get('scores_std') or {}
            std_scores = {k: float(v) for k, v in std.items()
                          if k in VALID_METRICS and v is not None}
            key = (model_id, bench_id)
            if key in results_by_key and std_scores:
                results_by_key[key]['scores_std'] = std_scores
                updated += 1
    print(f'Updated {updated} entries from {zippath.name}')


def update_from_excel(results_by_key, excel_path, sheet_map, method_map, model_id=None):
    import openpyxl
    updated = 0
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    for sheet_name, bench_id in sheet_map.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c).strip() if c else '' for c in rows[0]]
        header_to_metric = {}
        for idx, h in enumerate(headers):
            lower = h.lower()
            if lower in VALID_METRICS:
                header_to_metric[idx] = lower
        for row in rows[1:]:
            method = str(row[0]).strip() if row[0] is not None else ''
            if method == 'Raw (noisy)' or method == 'Raw (pseudo-deblended)':
                continue
            mid = method_map.get(method) if method_map else model_id
            if not mid:
                continue
            key = (mid, bench_id)
            if key not in results_by_key:
                continue
            std_scores = {}
            for idx, metric in header_to_metric.items():
                _, std = parse_mean_std(row[idx])
                if std is not None:
                    std_scores[metric] = std
            if std_scores:
                results_by_key[key]['scores_std'] = std_scores
                updated += 1
    print(f'Updated {updated} entries from {excel_path.name}')


def main():
    results = load_json(RESULTS_PATH)
    results_by_key = {(r['model_id'], r['benchmark_id']): r for r in results}
    total_before = sum(1 for r in results if r.get('scores_std'))
    print(f'Entries with scores_std before: {total_before}')

    update_from_interpolation_zip(results_by_key, SEGC3_INTERP_ZIP, SEGC3_MODEL_MAP, SEGC3_BENCH_MAP)
    update_from_interpolation_zip(results_by_key, AVO_INTERP_ZIP, AVO_MODEL_MAP, AVO_BENCH_MAP)

    update_from_excel(results_by_key, UNET_PP_EXCEL, UNET_PP_SHEET_MAP, None, model_id='zhou2018unet_plusplus_denoise')
    update_from_excel(results_by_key, GROUND_ROLL_EXCEL, GROUND_ROLL_SHEET_MAP, GROUND_ROLL_METHOD_MAP)
    update_from_excel(results_by_key, MULTIPLES_EXCEL, {'Multiples': 'multiples-attenuation'}, MULTIPLES_METHOD_MAP)

    total_after = sum(1 for r in results if r.get('scores_std'))
    print(f'Entries with scores_std after: {total_after}')

    save_json(RESULTS_PATH, results)
    print('Done')


if __name__ == '__main__':
    main()
