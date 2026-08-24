#!/usr/bin/env python3
"""Integrate latest interpolation results from JSON zips and Excels."""
import json
import re
import zipfile
import tempfile
from collections import Counter
from pathlib import Path

ROOT = Path(r'C:\Code\benchmark')
MODELS_PATH = ROOT / 'src' / 'data' / 'models.json'
BENCH_PATH = ROOT / 'src' / 'data' / 'benchmarks.json'
RESULTS_PATH = ROOT / 'src' / 'data' / 'results.json'

SEGC3_ZIP = Path(r'C:\Users\admin\Documents\WeChat Files\wxid_hvmr1h95e7jn22\FileStorage\File\2026-08\interpolation_json.zip')
AVO_ZIP = Path(r'C:\Users\admin\Documents\WeChat Files\wxid_hvmr1h95e7jn22\FileStorage\File\2026-08\avo_interpolation_json.zip')
SEGC3_EXCEL = Path(r'C:\Users\admin\Documents\WeChat Files\wxid_hvmr1h95e7jn22\FileStorage\File\2026-08\batch_evaluation.xlsx')
AVO_EXCEL = Path(r'C:\Users\admin\Documents\WeChat Files\wxid_hvmr1h95e7jn22\FileStorage\File\2026-08\batch_evaluation_avo.xlsx')

CORE_METRICS = ['snr', 'psnr', 'ssim', 'mae', 'mse', 'rmse']
BINNED_METRICS = [
    'eb_wse_medium_40_70_ne', 'eb_wse_medium_40_70_snr',
    'eb_wse_strong_70_100_ne', 'eb_wse_strong_70_100_snr',
    'eb_wse_very_weak_5_20_ne', 'eb_wse_very_weak_5_20_snr',
    'eb_wse_weak_20_40_ne', 'eb_wse_weak_20_40_snr',
    'fb_fre_high_ne', 'fb_fre_high_snr',
    'fb_fre_low_ne', 'fb_fre_low_snr',
    'fb_fre_mid_ne', 'fb_fre_mid_snr',
    'fb_fre_very_high_ne', 'fb_fre_very_high_snr',
]
VALID_METRICS = set(CORE_METRICS + BINNED_METRICS)
ALL_METRICS = CORE_METRICS + BINNED_METRICS

# Map raw model IDs in JSON to repository model IDs
SEGC3_MODEL_MAP = {
    'unet-interpolation': 'unet-interpolation',
    'unet-unpn-interpolation': 'unet-unpn-interpolation',
    'unet-plus-interpolation': 'unet-plus-interpolation',
    'attention-unet-interpolation': 'attention-unet-interpolation',
    'attention-unet-unpn-interpolation': 'attention-unet-unpn-interpolation',
    'attention-unet-plus-interpolation': 'attention-unet-plus-interpolation',
    'resunet-interpolation': 'res-unet-interpolation',
    'resunet-unpn-interpolation': 'resunet-unpn-interpolation',
    'resunet-plus-interpolation': 'resunet-plus-interpolation',
    'dncnn-interpolation': 'dncnn-interpolation',
    'dncnn-plus-interpolation': 'dncnn-plus-interpolation',
    'spnet-interpolation': 'spnet-interpolation',
}

# AVO model IDs are kept as-is
AVO_MODEL_MAP = {mid: mid for mid in [
    'avo-unet-interpolation',
    'avo-unet-plus-interpolation',
    'avo-attention-unet-interpolation',
    'avo-attention-unet-plus-interpolation',
    'avo-resunet-interpolation',
    'avo-resunet-plus-interpolation',
    'avo-dncnn-interpolation',
    'avo-dncnn-plus-interpolation',
    'avo-spnet-interpolation',
]}

SEGC3_BENCH_MAP = {
    'interp-random30': 'segc3-interp-random30',
    'interp-random50': 'segc3-interp-random50',
    'interp-random70': 'segc3-interp-random70',
    'interp-uniform50': 'segc3-interp-uniform50',
    'interp-uniform75': 'segc3-interp-uniform75',
    'interp-continuous20tr': 'segc3-interp-continuous20tr',
    'interp-continuous30tr': 'segc3-interp-continuous30tr',
    'interp-continuous40tr': 'segc3-interp-continuous40tr',
}

AVO_BENCH_MAP = {
    'avo-interp-random30': 'mobile-avo-interp-random30',
    'avo-interp-random50': 'mobile-avo-interp-random50',
    'avo-interp-random70': 'mobile-avo-interp-random70',
    'avo-interp-uniform50': 'mobile-avo-interp-uniform50',
    'avo-interp-uniform75': 'mobile-avo-interp-uniform75',
    'avo-interp-continuous20tr': 'mobile-avo-interp-continuous20tr',
    'avo-interp-continuous30tr': 'mobile-avo-interp-continuous30tr',
    'avo-interp-continuous40tr': 'mobile-avo-interp-continuous40tr',
}


def load_json(path):
    with path.open('r', encoding='utf-8') as f:
        return json.load(f)


def save_json(path, data):
    with path.open('w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write('\n')


def parse_mean(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s or s == '—':
        return None
    mean_str = re.split(r'[\s\xb1\u00b1]+', s)[0]
    try:
        return float(mean_str)
    except ValueError:
        return None


def extract_zip(zippath):
    tmp = tempfile.mkdtemp()
    with zipfile.ZipFile(zippath, 'r') as z:
        z.extractall(tmp)
    return Path(tmp)


def find_json_files(root):
    models, results = [], []
    for p in root.rglob('*.json'):
        if '__MACOSX' in p.parts:
            continue
        name = p.name.lower()
        if name.startswith('interpolation_result') and 'detailed' not in name:
            results.append(p)
        elif name.startswith('interpolation_model'):
            models.append(p)
    return models, results


def parse_params(excel_path):
    import openpyxl
    params = {}
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(c).strip() if c else '' for c in rows[0]]
        try:
            method_idx = headers.index('Method')
            param_idx = headers.index('Parameters (M)')
        except ValueError:
            continue
        for row in rows[1:]:
            method = str(row[method_idx]).strip() if row[method_idx] else ''
            val = row[param_idx]
            if method and val is not None:
                params[method] = float(val)
    return params


def build_params_by_model_id(excel_path, model_map):
    raw_params = parse_params(excel_path)
    # Map Excel method names to model IDs using model_map
    params_by_model = {}
    # Build reverse lookup from raw model ID to repo model ID
    for raw_mid, repo_mid in model_map.items():
        # Determine Excel method name from raw model id / known mapping
        method = None
        if raw_mid.startswith('avo-'):
            base = raw_mid[len('avo-'):].replace('-interpolation', '')
        else:
            base = raw_mid.replace('-interpolation', '')
        # Convert base to Excel method name
        base_to_method = {
            'unet': 'UNet',
            'unet-unpn': 'UNet (UNPN)',
            'unet-plus': 'UNet-Plus',
            'attention-unet': 'Attention UNet',
            'attention-unet-unpn': 'Attention UNet(UNPN)',
            'attention-unet-plus': 'Attention UNet-Plus',
            'resunet': 'ResUNet',
            'resunet-unpn': 'ResUNet(UNPN)',
            'resunet-plus': 'ResUNet-Plus',
            'dncnn': 'DnCNN',
            'dncnn-plus': 'DnCNN-Plus',
            'spnet': 'SPNet',
        }
        method = base_to_method.get(base)
        if method and method in raw_params:
            params_by_model[repo_mid] = raw_params[method]
    return params_by_model


def add_models(models_data, model_files, model_map, params_by_model):
    existing_ids = {m['id'] for m in models_data}
    added = 0
    for mp in model_files:
        raw = load_json(mp)
        raw_id = raw.get('id')
        if not raw_id or raw_id not in model_map:
            continue
        mid = model_map[raw_id]
        if mid in existing_ids:
            # update parameters if missing
            for m in models_data:
                if m['id'] == mid and m.get('parameters_m') is None and mid in params_by_model:
                    m['parameters_m'] = params_by_model[mid]
            continue
        model = {
            'id': mid,
            'name': raw.get('name') or mid,
            'authors': raw.get('authors') or 'Unknown',
            'org': raw.get('org') or 'Unknown',
            'year': raw.get('year') or 2026,
            'emoji': raw.get('emoji') or '🔧',
            'type': raw.get('type', 'deep_learning'),
            'tasks': raw.get('tasks', ['interpolation']),
            'description': raw.get('description') or f"{raw.get('name') or mid} for seismic interpolation.",
            'paper_url': raw.get('paper_url') or None,
            'code_url': raw.get('code_url') or None,
            'weights_url': raw.get('weights_url') or None,
            'is_open_source': True,
            'parameters_m': params_by_model.get(mid),
        }
        models_data.append(model)
        existing_ids.add(mid)
        added += 1
        print(f'Added model {mid}')
    return added


def integrate_results(results_data, result_files, model_map, bench_map):
    existing = {(r['model_id'], r['benchmark_id']): r for r in results_data}
    added = 0
    replaced = 0
    skipped = 0
    for rp in result_files:
        entries = load_json(rp)
        if not isinstance(entries, list):
            entries = entries.get('results', [entries])
        for entry in entries:
            raw_bench = entry.get('benchmark_id')
            raw_model = entry.get('model_id')
            if raw_bench not in bench_map or raw_model not in model_map:
                skipped += 1
                continue
            bench_id = bench_map[raw_bench]
            model_id = model_map[raw_model]
            scores = {k: float(v) for k, v in entry.get('scores', {}).items()
                      if k in VALID_METRICS and v is not None}
            result = {
                'model_id': model_id,
                'benchmark_id': bench_id,
                'scores': scores,
                'paper_url': entry.get('paper_url') or None,
                'code_url': entry.get('code_url') or None,
                'date_added': '2026-08-24',
            }
            key = (model_id, bench_id)
            if key in existing:
                existing[key].update(result)
                replaced += 1
            else:
                results_data.append(result)
                existing[key] = result
                added += 1
    print(f'Results: added={added}, replaced={replaced}, skipped={skipped}')
    return added, replaced


def find_index(benchmarks, predicate):
    for i, b in enumerate(benchmarks):
        if predicate(b):
            return i
    return len(benchmarks)


def create_missing_benchmarks(benchmarks):
    new_specs = [
        # SEGC3
        ('segc3-interp-uniform75', 'SEGC3 Uniform Missing 75%', 'SEGC3 Uniform Missing', '75% uniformly missing traces', 'Uniform Missing',
         ['datasets/segc3-interp-uniform-raw.png', 'datasets/segc3-interp-uniform-label.png']),
        # Mobile AVO
        ('mobile-avo-interp-random70', 'Mobile AVO Random Missing 70%', 'Mobile AVO Random Missing', '70% randomly missing traces', 'Random Missing',
         ['datasets/mobile-avo-data.png']),
        ('mobile-avo-interp-uniform75', 'Mobile AVO Uniform Missing 75%', 'Mobile AVO Uniform Missing', '75% uniformly missing traces', 'Uniform Missing',
         ['datasets/mobile-avo-data.png']),
    ]
    created = []
    for bid, name, group, desc_suffix, tag, gallery in new_specs:
        if any(b['id'] == bid for b in benchmarks):
            print(f'Benchmark {bid} already exists')
            continue
        # find template
        if bid.startswith('segc3'):
            tmpl_id = 'segc3-interp-uniform70'
        elif 'random' in bid:
            tmpl_id = 'mobile-avo-interp-random50'
        else:
            tmpl_id = 'mobile-avo-interp-uniform50'
        tmpl = next(b for b in benchmarks if b['id'] == tmpl_id)
        bench = {
            'id': bid,
            'name': name,
            'group_name': group,
            'dataset_name': tmpl['dataset_name'],
            'task': 'interpolation',
            'icon': tmpl['icon'],
            'description': f"{tmpl['dataset_name']} interpolation benchmark with {desc_suffix}.",
            'data_source': tmpl['data_source'],
            'dimensions': tmpl['dimensions'],
            'primary_metric': 'snr',
            'metrics': ALL_METRICS,
            'tags': [tag if t == 'Interpolation' else t for t in tmpl['tags']],
            'citation': tmpl['citation'],
            'download_url': tmpl['download_url'],
            'model_count': 0,
            'gallery': gallery,
        }
        # insert near template
        idx = find_index(benchmarks, lambda b: b['id'] == tmpl_id) + 1
        benchmarks.insert(idx, bench)
        created.append(bid)
        print(f'Added benchmark {bid}')
    return created


def recalc_counts(benchmarks, results):
    counts = Counter(r['benchmark_id'] for r in results)
    for b in benchmarks:
        if b['task'] == 'interpolation':
            old = b['model_count']
            b['model_count'] = counts.get(b['id'], 0)
            if old != b['model_count']:
                print(f'Recalc {b["id"]}: {old} -> {b["model_count"]}')


def main():
    models = load_json(MODELS_PATH)
    benchmarks = load_json(BENCH_PATH)
    results = load_json(RESULTS_PATH)

    segc3_params = build_params_by_model_id(SEGC3_EXCEL, SEGC3_MODEL_MAP)
    avo_params = build_params_by_model_id(AVO_EXCEL, AVO_MODEL_MAP)

    # SEGC3
    tmp = extract_zip(SEGC3_ZIP)
    root = tmp / 'interpolation_json'
    segc3_models, segc3_results = find_json_files(root)
    add_models(models, segc3_models, SEGC3_MODEL_MAP, segc3_params)
    integrate_results(results, segc3_results, SEGC3_MODEL_MAP, SEGC3_BENCH_MAP)

    # Mobile AVO
    tmp = extract_zip(AVO_ZIP)
    root = tmp / 'avo_interpolation_json'
    avo_models, avo_results = find_json_files(root)
    add_models(models, avo_models, AVO_MODEL_MAP, avo_params)
    integrate_results(results, avo_results, AVO_MODEL_MAP, AVO_BENCH_MAP)

    create_missing_benchmarks(benchmarks)
    recalc_counts(benchmarks, results)

    save_json(MODELS_PATH, models)
    save_json(BENCH_PATH, benchmarks)
    save_json(RESULTS_PATH, results)
    print('Done')


if __name__ == '__main__':
    main()
