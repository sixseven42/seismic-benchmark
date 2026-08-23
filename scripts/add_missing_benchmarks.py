#!/usr/bin/env python3
"""Add missing interpolation benchmark, Mobile AVO random-noise benchmarks, and integrate UNet++ AVO results."""
import json
import re
import zipfile
import tempfile
from collections import Counter
from pathlib import Path

ROOT = Path(r'C:\Code\benchmark')
MODELS_PATH = ROOT / 'src' / 'data' / 'models.json'
BENCH_PATH = ROOT / 'src' / 'data' / 'benchmarks.json'
RESULTS_PATH = ROOT / 'src' / 'data' / 'results.json'

UNET_PP_EXCEL = Path(r'C:\论文\SeisBench\seismic_bench可视化\random\batch_evaluation_unet_plusplus.xlsx')
SYN_ZIP = Path(r'C:\Users\admin\Documents\WeChat Files\wxid_hvmr1h95e7jn22\FileStorage\File\2026-08\interp_syn_czt0822.zip')

CORE_METRICS = ['snr', 'psnr', 'ssim', 'mae', 'mse', 'rmse']
BINNED_METRICS = [
    'eb_wse_medium_40_70_ne', 'eb_wse_medium_40_70_snr',
    'eb_wse_strong_70_100_ne', 'eb_wse_strong_70_100_snr',
    'eb_wse_very_weak_5_20_ne', 'eb_wse_very_weak_5_20_snr',
    'eb_wse_weak_20_40_ne', 'eb_wse_weak_20_40_snr',
    'fb_fre_high_ne', 'fb_fre_high_snr',
    'fb_fre_low_ne', 'fb_fre_low_snr',
    'fb_fre_mid_ne', 'fb_fre_mid_snr',
    'fb_fre_very_high_ne', 'fb_fre_very_high_snr',
]
VALID_METRICS = set(CORE_METRICS + BINNED_METRICS)
ALL_METRICS = CORE_METRICS + BINNED_METRICS


def load_json(path):
    with path.open('r', encoding='utf-8') as f:
        return json.load(f)


def save_json(path, data):
    with path.open('w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write('\n')


def parse_mean(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s or s == '—':
        return None
    mean_str = re.split(r'[\s\xb1\u00b1]+', s)[0]
    try:
        return float(mean_str)
    except ValueError:
        return None


def find_insert_index(benchmarks, predicate):
    for i, b in enumerate(benchmarks):
        if predicate(b):
            return i
    return len(benchmarks)


def add_interpolation_benchmark(benchmarks):
    if any(b['id'] == 'segc3-interp-random10-30' for b in benchmarks):
        print('Benchmark segc3-interp-random10-30 already exists')
        return
    template = next(b for b in benchmarks if b['id'] == 'segc3-interp-random30')
    new_bench = {
        'id': 'segc3-interp-random10-30',
        'name': 'SEGC3 Random Missing 10-30%',
        'group_name': template['group_name'],
        'dataset_name': template['dataset_name'],
        'task': 'interpolation',
        'icon': template['icon'],
        'description': 'SEG C3 seismic interpolation benchmark with 10-30% randomly missing traces.',
        'data_source': template['data_source'],
        'dimensions': template['dimensions'],
        'primary_metric': 'snr',
        'metrics': ALL_METRICS,
        'tags': ['SEG C3', 'Interpolation', 'Random Missing', 'Synthetic'],
        'citation': template['citation'],
        'download_url': template['download_url'],
        'model_count': 0,
        'gallery': template.get('gallery', []),
    }
    idx = find_insert_index(benchmarks, lambda b: b['id'] == 'segc3-interp-random30')
    benchmarks.insert(idx + 1, new_bench)
    print('Added benchmark segc3-interp-random10-30')


def add_mobile_avo_benchmarks(benchmarks):
    variants = [
        ('gaussian', 'snrneg5', '-5 dB', 'Gaussian'),
        ('gaussian', 'snr0', '0 dB', 'Gaussian'),
        ('gaussian', 'snr5', '+5 dB', 'Gaussian'),
        ('poisson', 'snrneg5', '-5 dB', 'Poisson'),
        ('poisson', 'snr0', '0 dB', 'Poisson'),
        ('poisson', 'snr5', '+5 dB', 'Poisson'),
    ]
    added = 0
    insert_idx = find_insert_index(benchmarks, lambda b: b['id'] == 'fbp-geomseg-all')
    for noise, suffix, snr_label, noise_label in variants:
        bid = f'mobile-avo-random-noise-{noise}-{suffix}'
        if any(b['id'] == bid for b in benchmarks):
            print(f'Benchmark {bid} already exists')
            continue
        bench = {
            'id': bid,
            'name': f'Mobile AVO Random Noise {noise_label} SNR {snr_label}',
            'group_name': 'Mobile AVO Random Noise',
            'dataset_name': 'Mobile AVO Random Noise',
            'task': 'random_noise_suppression',
            'icon': '🌊',
            'description': f'Mobile AVO field random-noise attenuation benchmark with {noise_label} noise at {snr_label}.',
            'data_source': 'field',
            'dimensions': 'Mobile AVO field gathers',
            'primary_metric': 'snr',
            'metrics': ALL_METRICS,
            'tags': ['Mobile AVO', 'Random Noise', noise_label, 'Field'],
            'citation': '',
            'download_url': '',
            'model_count': 0,
            'gallery': [],
        }
        benchmarks.insert(insert_idx, bench)
        insert_idx += 1
        added += 1
        print(f'Added benchmark {bid}')
    return added


def add_missing_yu_result(results):
    if any(r['model_id'] == 'yu2022_anet_interpolation' and r['benchmark_id'] == 'segc3-interp-random10-30' for r in results):
        print('yu2022_anet_interpolation result for segc3-interp-random10-30 already exists')
        return
    tmp = tempfile.mkdtemp()
    with zipfile.ZipFile(SYN_ZIP, 'r') as z:
        z.extractall(tmp)
    yu_path = Path(tmp) / 'interp_syn_czt0822' / 'interpolation_result_yu2022_anet.json'
    entries = load_json(yu_path)
    for entry in entries:
        if entry.get('benchmark_id') == 'interp-random-10-30':
            scores = {k: float(v) for k, v in entry['scores'].items() if k in VALID_METRICS and v is not None}
            results.append({
                'model_id': 'yu2022_anet_interpolation',
                'benchmark_id': 'segc3-interp-random10-30',
                'scores': scores,
                'paper_url': entry.get('paper_url'),
                'code_url': entry.get('code_url'),
                'date_added': '2026-08-23',
            })
            print('Added yu2022_anet_interpolation result for segc3-interp-random10-30')
            return
    print('Warning: interp-random-10-30 not found in yu2022 result file')


def integrate_unet_pp_avo(results):
    import openpyxl
    wb = openpyxl.load_workbook(UNET_PP_EXCEL, data_only=True)
    sheet_map = {
        'RN-AVO Gaussian -5dB': 'mobile-avo-random-noise-gaussian-snrneg5',
        'RN-AVO Gaussian +0dB': 'mobile-avo-random-noise-gaussian-snr0',
        'RN-AVO Gaussian +5dB': 'mobile-avo-random-noise-gaussian-snr5',
        'RN-AVO Poisson -5dB': 'mobile-avo-random-noise-poisson-snrneg5',
        'RN-AVO Poisson +0dB': 'mobile-avo-random-noise-poisson-snr0',
        'RN-AVO Poisson +5dB': 'mobile-avo-random-noise-poisson-snr5',
    }
    model_id = 'zhou2018unet_plusplus_denoise'
    existing = {(r['model_id'], r['benchmark_id']): r for r in results}
    added = 0
    for sheet_name, bench_id in sheet_map.items():
        ws = wb[sheet_name]
        headers = [str(c).strip() if c else '' for c in next(ws.iter_rows(min_row=1, values_only=True))]
        header_to_metric = {i: h.lower() for i, h in enumerate(headers) if h.lower() in VALID_METRICS}
        for row in ws.iter_rows(min_row=2, values_only=True):
            method = str(row[0]).strip() if row[0] else ''
            if method != 'UNet-Plus':
                continue
            scores = {}
            for idx, metric in header_to_metric.items():
                val = parse_mean(row[idx])
                if val is not None:
                    scores[metric] = val
            result = {
                'model_id': model_id,
                'benchmark_id': bench_id,
                'scores': scores,
                'paper_url': 'https://arxiv.org/abs/1807.10165',
                'code_url': 'https://github.com/sixseven42/seismic-benchmark-code/blob/main/model/random_noise_suppression/unet_plusplus.py',
                'date_added': '2026-08-23',
            }
            key = (model_id, bench_id)
            if key in existing:
                existing[key].update(result)
                print(f'Replaced result {model_id} / {bench_id}')
            else:
                results.append(result)
                existing[key] = result
                added += 1
                print(f'Added result {model_id} / {bench_id}')
            break
    return added


def update_model_counts(benchmarks, results):
    counts = Counter(r['benchmark_id'] for r in results)
    affected = {'segc3-interp-random10-30'}
    affected.update(b['id'] for b in benchmarks if b['id'].startswith('mobile-avo-random-noise'))
    for b in benchmarks:
        if b['id'] in affected:
            old = b['model_count']
            b['model_count'] = counts.get(b['id'], 0)
            print(f'Updated {b["id"]} model_count: {old} -> {b["model_count"]}')


def main():
    models = load_json(MODELS_PATH)
    benchmarks = load_json(BENCH_PATH)
    results = load_json(RESULTS_PATH)

    add_interpolation_benchmark(benchmarks)
    add_mobile_avo_benchmarks(benchmarks)
    add_missing_yu_result(results)
    integrate_unet_pp_avo(results)
    update_model_counts(benchmarks, results)

    save_json(MODELS_PATH, models)
    save_json(BENCH_PATH, benchmarks)
    save_json(RESULTS_PATH, results)
    print('Done')


if __name__ == '__main__':
    main()
