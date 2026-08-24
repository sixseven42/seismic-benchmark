#!/usr/bin/env python3
"""Integrate UNet++ deblending results from batch_evaluation_unet_plusplus.xlsx."""
import json
import re
from pathlib import Path

ROOT = Path(r'C:\Code\benchmark')
EXCEL = Path(r'C:\论文\SeisBench\seismic_bench可视化\random\batch_evaluation_unet_plusplus.xlsx')
MODELS_PATH = ROOT / 'src' / 'data' / 'models.json'
BENCH_PATH = ROOT / 'src' / 'data' / 'benchmarks.json'
RESULTS_PATH = ROOT / 'src' / 'data' / 'results.json'
PAGE_PATH = ROOT / 'src' / 'pages' / 'BenchmarksPage.tsx'

SHEET_MAP = {
    'Deblending-AVO T03 mod': 'mobile-avo-deblending-t03-mod',
    'Deblending T02 mod': 'segc3-deblending-t02-mod',
    'Deblending T02 simp': 'segc3-deblending-t02-simp',
    'Deblending T02 comp': 'segc3-deblending-t02-comp',
}

BENCH_SPECS = {
    'mobile-avo-deblending-t03-mod': {
        'name': 'Mobile AVO Deblending T03 Mod',
        'group_name': 'Mobile AVO Deblending',
        'dataset_name': 'Mobile AVO Deblending',
        'description': 'Mobile AVO Viking Graben Line 12 open-source 2D marine field dataset deblending benchmark with T03 mod blending scenario.',
        'data_source': 'field',
        'dimensions': 'Mobile AVO field gathers',
        'tags': ['Mobile AVO', 'Deblending', 'Field'],
        'citation': 'https://wiki.seg.org/wiki/Mobil_AVO_viking_graben_line_12',
        'download_url': 'https://wiki.seg.org/wiki/Mobil_AVO_viking_graben_line_12',
        'gallery': ['datasets/mobile-avo-data.png'],
    },
    'segc3-deblending-t02-mod': {
        'name': 'SEGC3 Deblending T02 Mod',
        'group_name': 'SEGC3 Deblending',
        'dataset_name': 'SEG C3 Deblending',
        'description': 'SEG C3 synthetic deblending benchmark with T02 mod blending scenario.',
        'data_source': 'synthetic',
        'dimensions': '9 × 201 × 625 (shots, traces, time)',
        'tags': ['SEG C3', 'Deblending', 'Synthetic'],
        'citation': 'https://wiki.seg.org/wiki/C3',
        'download_url': 'https://wiki.seg.org/wiki/C3',
        'gallery': ['datasets/segc3-raw.png', 'datasets/segc3-label.png'],
    },
    'segc3-deblending-t02-simp': {
        'name': 'SEGC3 Deblending T02 Simp',
        'group_name': 'SEGC3 Deblending',
        'dataset_name': 'SEG C3 Deblending',
        'description': 'SEG C3 synthetic deblending benchmark with T02 simp blending scenario.',
        'data_source': 'synthetic',
        'dimensions': '9 × 201 × 625 (shots, traces, time)',
        'tags': ['SEG C3', 'Deblending', 'Synthetic'],
        'citation': 'https://wiki.seg.org/wiki/C3',
        'download_url': 'https://wiki.seg.org/wiki/C3',
        'gallery': ['datasets/segc3-raw.png', 'datasets/segc3-label.png'],
    },
    'segc3-deblending-t02-comp': {
        'name': 'SEGC3 Deblending T02 Comp',
        'group_name': 'SEGC3 Deblending',
        'dataset_name': 'SEG C3 Deblending',
        'description': 'SEG C3 synthetic deblending benchmark with T02 comp blending scenario.',
        'data_source': 'synthetic',
        'dimensions': '9 × 201 × 625 (shots, traces, time)',
        'tags': ['SEG C3', 'Deblending', 'Synthetic'],
        'citation': 'https://wiki.seg.org/wiki/C3',
        'download_url': 'https://wiki.seg.org/wiki/C3',
        'gallery': ['datasets/segc3-raw.png', 'datasets/segc3-label.png'],
    },
}

CORE_METRICS = ['snr', 'psnr', 'ssim', 'mae', 'mse', 'rmse']
BINNED_METRICS = [
    'eb_wse_medium_40_70_ne', 'eb_wse_medium_40_70_snr',
    'eb_wse_strong_70_100_ne', 'eb_wse_strong_70_100_snr',
    'eb_wse_very_weak_5_20_ne', 'eb_wse_very_weak_5_20_snr',
    'eb_wse_weak_20_40_ne', 'eb_wse_weak_20_40_snr',
    'fb_fre_high_ne', 'fb_fre_high_snr',
    'fb_fre_low_ne', 'fb_fre_low_snr',
    'fb_fre_mid_ne', 'fb_fre_mid_snr',
    'fb_fre_very_high_ne', 'fb_fre_very_high_snr',
]
ALL_METRICS = CORE_METRICS + BINNED_METRICS

MODEL_ID = 'zhou2018unet_plusplus_denoise'


def load_json(path):
    with path.open('r', encoding='utf-8') as f:
        return json.load(f)


def save_json(path, data):
    with path.open('w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write('\n')


def parse_mean(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s or s == '—':
        return None
    mean_str = re.split(r'[\s\xb1\u00b1]+', s)[0]
    try:
        return float(mean_str)
    except ValueError:
        return None


def update_model(models):
    for m in models:
        if m['id'] == MODEL_ID:
            tasks = set(m.get('tasks', []))
            tasks.add('deblending')
            m['tasks'] = list(tasks)
            if 'deblending' not in m.get('description', '').lower():
                m['description'] = m['description'].rstrip('.') + ' and seismic deblending.'
            print(f'Updated model {MODEL_ID}: tasks={m["tasks"]}')
            return
    raise ValueError(f'Model {MODEL_ID} not found')


def find_index(benchmarks, predicate):
    for i, b in enumerate(benchmarks):
        if predicate(b):
            return i
    return len(benchmarks)


def create_benchmarks(benchmarks):
    insert_idx = find_index(benchmarks, lambda b: b['task'] == 'first_arrival_picking')
    added = 0
    for bid, spec in BENCH_SPECS.items():
        if any(b['id'] == bid for b in benchmarks):
            print(f'Benchmark {bid} already exists')
            continue
        bench = {
            'id': bid,
            'name': spec['name'],
            'group_name': spec['group_name'],
            'dataset_name': spec['dataset_name'],
            'task': 'deblending',
            'icon': '🌀',
            'description': spec['description'],
            'data_source': spec['data_source'],
            'dimensions': spec['dimensions'],
            'primary_metric': 'snr',
            'metrics': ALL_METRICS,
            'tags': spec['tags'],
            'citation': spec['citation'],
            'download_url': spec['download_url'],
            'model_count': 0,
            'gallery': spec['gallery'],
        }
        benchmarks.insert(insert_idx, bench)
        insert_idx += 1
        added += 1
        print(f'Added benchmark {bid}')
    return added


def integrate_results(results):
    import openpyxl
    existing = {(r['model_id'], r['benchmark_id']): r for r in results}
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    added = 0
    replaced = 0
    for sheet, bench_id in SHEET_MAP.items():
        ws = wb[sheet]
        headers = [str(c).strip() if c is not None else '' for c in next(ws.iter_rows(min_row=1, values_only=True))]
        header_to_metric = {}
        for idx, h in enumerate(headers):
            lower = h.lower()
            if lower in ALL_METRICS:
                header_to_metric[idx] = lower

        for row in ws.iter_rows(min_row=2, values_only=True):
            method = str(row[0]).strip() if row[0] is not None else ''
            if method != 'UNet-Plus':
                continue
            scores = {}
            for idx, metric in header_to_metric.items():
                val = parse_mean(row[idx])
                if val is not None:
                    scores[metric] = val
            key = (MODEL_ID, bench_id)
            result = {
                'model_id': MODEL_ID,
                'benchmark_id': bench_id,
                'scores': scores,
                'paper_url': 'https://arxiv.org/abs/1807.10165',
                'code_url': 'https://github.com/sixseven42/seismic-benchmark-code/blob/main/model/random_noise_suppression/unet_plusplus.py',
                'date_added': '2026-08-24',
            }
            if key in existing:
                existing[key].update(result)
                replaced += 1
                print(f'Replaced result {MODEL_ID} / {bench_id}')
            else:
                results.append(result)
                existing[key] = result
                added += 1
                print(f'Added result {MODEL_ID} / {bench_id}')
            break
    print(f'Results: added={added}, replaced={replaced}')


def update_group_descriptions(page_text):
    # Insert before the closing brace of GROUP_DESCRIPTIONS
    new_entries = (
        "  'SEGC3 Deblending': 'Synthetic 3D seismic deblending benchmarks based on the SEG China 3D (SEGC3) geological model. Each variant uses a different T02 blending scenario (mod, simp, comp), providing blended inputs and corresponding labels for evaluating methods that separate overlapping shot or blended arrivals.',\n"
        "  'Mobile AVO Deblending': 'A 2D marine field deblending benchmark using the Mobile AVO Viking Graben Line 12 open-source dataset. The T03 mod scenario evaluates deblending methods on real marine seismic data with overlapping arrivals.',\n"
    )
    # Find the line with the closing brace of the object literal
    pattern = r"(const GROUP_DESCRIPTIONS: Record<string, string> = \{[\s\S]*?)\n\};"
    match = re.search(pattern, page_text)
    if match:
        insert_pos = match.end(1)
        page_text = page_text[:insert_pos] + '\n' + new_entries + page_text[insert_pos:]
        print('Updated BenchmarksPage.tsx group descriptions')
    else:
        print('Warning: could not find GROUP_DESCRIPTIONS block')
    return page_text


def recalc_counts(benchmarks, results):
    from collections import Counter
    counts = Counter(r['benchmark_id'] for r in results)
    for b in benchmarks:
        if b['task'] == 'deblending':
            old = b['model_count']
            b['model_count'] = counts.get(b['id'], 0)
            if old != b['model_count']:
                print(f'Recalc {b["id"]}: {old} -> {b["model_count"]}')


def main():
    models = load_json(MODELS_PATH)
    benchmarks = load_json(BENCH_PATH)
    results = load_json(RESULTS_PATH)

    update_model(models)
    create_benchmarks(benchmarks)
    integrate_results(results)
    recalc_counts(benchmarks, results)

    save_json(MODELS_PATH, models)
    save_json(BENCH_PATH, benchmarks)
    save_json(RESULTS_PATH, results)

    page_text = PAGE_PATH.read_text(encoding='utf-8')
    page_text = update_group_descriptions(page_text)
    PAGE_PATH.write_text(page_text, encoding='utf-8')

    print('Done')


if __name__ == '__main__':
    main()
