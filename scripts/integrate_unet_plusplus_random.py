#!/usr/bin/env python3
"""Integrate UNet++ random-noise SEGC3 results into website data."""
import json
import re
from pathlib import Path

ROOT = Path(r'C:\Code\benchmark')
EXCEL = Path(r'C:\论文\SeisBench\seismic_bench可视化\random\batch_evaluation_unet_plusplus.xlsx')
MODEL_JSON = Path(r'C:\论文\SeisBench\seismic_bench可视化\random\unet_plusplus_model.json')

MODELS_PATH = ROOT / 'src' / 'data' / 'models.json'
BENCH_PATH = ROOT / 'src' / 'data' / 'benchmarks.json'
RESULTS_PATH = ROOT / 'src' / 'data' / 'results.json'

SHEET_MAP = {
    'RN-SEGC3 Gaussian -5dB': 'segc3-random-noise-gaussian-snrneg5',
    'RN-SEGC3 Gaussian +0dB': 'segc3-random-noise-gaussian-snr0',
    'RN-SEGC3 Gaussian +5dB': 'segc3-random-noise-gaussian-snr5',
    'RN-SEGC3 Poisson -5dB': 'segc3-random-noise-poisson-snrneg5',
    'RN-SEGC3 Poisson +0dB': 'segc3-random-noise-poisson-snr0',
    'RN-SEGC3 Poisson +5dB': 'segc3-random-noise-poisson-snr5',
}

CORE_METRICS = ['snr', 'psnr', 'ssim', 'mae', 'mse', 'rmse']
BINNED_METRICS = [
    'eb_wse_medium_40_70_ne',
    'eb_wse_medium_40_70_snr',
    'eb_wse_strong_70_100_ne',
    'eb_wse_strong_70_100_snr',
    'eb_wse_very_weak_5_20_ne',
    'eb_wse_very_weak_5_20_snr',
    'eb_wse_weak_20_40_ne',
    'eb_wse_weak_20_40_snr',
    'fb_fre_high_ne',
    'fb_fre_high_snr',
    'fb_fre_low_ne',
    'fb_fre_low_snr',
    'fb_fre_mid_ne',
    'fb_fre_mid_snr',
    'fb_fre_very_high_ne',
    'fb_fre_very_high_snr',
]
ALL_METRICS = CORE_METRICS + BINNED_METRICS


def parse_mean(value):
    if value is None:
        return None
    s = str(value).strip()
    if not s or s == '—':
        return None
    # split on plus-minus (U+00B1) and take mean
    mean_str = re.split(r'[\s\xb1\u00b1]+', s)[0]
    try:
        return float(mean_str)
    except ValueError:
        return None


def main():
    import openpyxl

    # 1. Add model if missing
    with MODELS_PATH.open('r', encoding='utf-8') as f:
        models = json.load(f)

    model_id = 'zhou2018unet_plusplus_denoise'
    if not any(m.get('id') == model_id for m in models):
        raw_model = json.loads(MODEL_JSON.read_text(encoding='utf-8'))
        # Fill required fields and keep only random_noise_suppression (no deblending benchmarks exist yet)
        model = {
            'id': raw_model['id'],
            'name': raw_model['name'],
            'authors': raw_model['authors'],
            'org': 'Arizona State University',
            'year': raw_model['year'],
            'emoji': '🔀',
            'type': raw_model.get('type', 'deep_learning'),
            'tasks': ['random_noise_suppression'],
            'description': 'Nested U-Net++ architecture with dense skip connections for seismic random-noise suppression.',
            'paper_url': raw_model.get('paper_url'),
            'code_url': 'https://github.com/sixseven42/seismic-benchmark-code/blob/main/model/random_noise_suppression/unet_plusplus.py',
            'weights_url': None,
            'is_open_source': True,
            'parameters_m': raw_model.get('parameters_m'),
        }
        models.append(model)
        print(f'Added model {model_id}')
    else:
        print(f'Model {model_id} already exists')

    with MODELS_PATH.open('w', encoding='utf-8') as f:
        json.dump(models, f, ensure_ascii=False, indent=2)
        f.write('\n')

    # 2. Update benchmarks
    with BENCH_PATH.open('r', encoding='utf-8') as f:
        benchmarks = json.load(f)

    bench_by_id = {b['id']: b for b in benchmarks}
    for bench_id in SHEET_MAP.values():
        b = bench_by_id[bench_id]
        b['metrics'] = ALL_METRICS
        b['model_count'] = b.get('model_count', 0) + 1
        print(f'Updated benchmark {bench_id}: model_count={b["model_count"]}')

    with BENCH_PATH.open('w', encoding='utf-8') as f:
        json.dump(benchmarks, f, ensure_ascii=False, indent=2)
        f.write('\n')

    # 3. Extract results and append
    with RESULTS_PATH.open('r', encoding='utf-8') as f:
        results = json.load(f)

    existing_keys = {(r['model_id'], r['benchmark_id']) for r in results}
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    added = 0
    for sheet, bench_id in SHEET_MAP.items():
        ws = wb[sheet]
        headers = [str(c).strip() if c is not None else '' for c in next(ws.iter_rows(min_row=1, values_only=True))]
        header_to_metric = {}
        for idx, h in enumerate(headers):
            lower = h.lower()
            if lower in ALL_METRICS:
                header_to_metric[idx] = lower
            elif lower.endswith('_std'):
                pass
            elif 'energy_ratio' in lower or 'frequency_range_hz' in lower:
                pass
            else:
                pass

        for row in ws.iter_rows(min_row=2, values_only=True):
            method = str(row[0]).strip() if row[0] is not None else ''
            if method != 'UNet-Plus':
                continue
            scores = {}
            for idx, metric in header_to_metric.items():
                val = parse_mean(row[idx])
                if val is not None:
                    scores[metric] = val
            result_key = (model_id, bench_id)
            if result_key in existing_keys:
                # replace existing
                for r in results:
                    if r['model_id'] == model_id and r['benchmark_id'] == bench_id:
                        r['scores'] = scores
                        break
                print(f'Replaced result {model_id} / {bench_id}')
            else:
                results.append({
                    'model_id': model_id,
                    'benchmark_id': bench_id,
                    'scores': scores,
                    'paper_url': 'https://arxiv.org/abs/1807.10165',
                    'code_url': 'https://github.com/sixseven42/seismic-benchmark-code/blob/main/model/random_noise_suppression/unet_plusplus.py',
                    'date_added': '2026-08-23',
                })
                existing_keys.add(result_key)
                added += 1
                print(f'Added result {model_id} / {bench_id}')
            break

    with RESULTS_PATH.open('w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
        f.write('\n')

    print(f'Done. Added {added} new result entries.')


if __name__ == '__main__':
    main()
